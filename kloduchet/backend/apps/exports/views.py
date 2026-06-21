import io

import openpyxl
from django.http import HttpResponse
from rest_framework.views import APIView

from apps.analytics.views import (
    GROSS_SALES_SUM,
    RETURNS_SUM,
    SALES_VS_PAYMENTS_DISCLAIMER,
    TRUNC_BY_GROUP,
    _filtered_payments_queryset,
    _filtered_queryset,
)
from django.db.models import Sum, Count


class SalesTimelineExportView(APIView):
    def get(self, request):
        group = request.query_params.get("group", "month")
        trunc = TRUNC_BY_GROUP.get(group, TRUNC_BY_GROUP["month"])

        qs = _filtered_queryset(request)
        rows = (
            qs.annotate(period=trunc("sale_date"))
            .values("period")
            .annotate(
                gross_sales_total=GROSS_SALES_SUM,
                returns_total=RETURNS_SUM,
                amount_total=Sum("amount"),
                quantity_total=Sum("quantity"),
                documents_count=Count("sales_doc_number", distinct=True),
            )
            .order_by("period")
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Продажи по времени"
        sheet.append(
            ["Период", "Сумма продаж", "Возвраты и корректировки", "Чистая сумма", "Количество", "Документов"]
        )

        for row in rows:
            if row["period"] is None:
                continue
            sheet.append(
                [
                    row["period"],
                    float(row["gross_sales_total"] or 0),
                    float(row["returns_total"] or 0),
                    float(row["amount_total"] or 0),
                    float(row["quantity_total"] or 0),
                    row["documents_count"] or 0,
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sales_timeline.xlsx"'
        return response


class SalesVsPaymentsExportView(APIView):
    def get(self, request):
        sales_qs = _filtered_queryset(request)
        payments_qs = _filtered_payments_queryset(request)

        sales_rows = sales_qs.values("organization_id", "legal_entity", "contract_number").annotate(
            sales_total=Sum("amount")
        )
        payments_rows = payments_qs.values("organization_id", "legal_entity", "contract_number").annotate(
            payments_total=Sum("amount")
        )

        combined = {}
        for row in sales_rows:
            key = (row["organization_id"], row["legal_entity"], row["contract_number"])
            combined[key] = {**row, "sales_total": row["sales_total"] or 0, "payments_total": 0}
        for row in payments_rows:
            key = (row["organization_id"], row["legal_entity"], row["contract_number"])
            entry = combined.setdefault(key, {**row, "sales_total": 0, "payments_total": 0})
            entry["payments_total"] = row["payments_total"] or 0

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Продажи и оплаты"
        sheet.append(["Юр. лицо", "Договор", "Продажи", "Оплаты", "Разница"])
        for entry in sorted(combined.values(), key=lambda r: r["sales_total"], reverse=True):
            sheet.append(
                [
                    entry["legal_entity"],
                    entry["contract_number"],
                    float(entry["sales_total"]),
                    float(entry["payments_total"]),
                    float(entry["sales_total"]) - float(entry["payments_total"]),
                ]
            )
        sheet.append([])
        sheet.append([SALES_VS_PAYMENTS_DISCLAIMER])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sales_vs_payments.xlsx"'
        return response


class SalesProductsExportView(APIView):
    def get(self, request):
        qs = _filtered_queryset(request)
        total_amount = qs.aggregate(total=Sum("amount"))["total"] or 0
        rows = (
            qs.values("nomenclature")
            .annotate(
                gross_sales_total=GROSS_SALES_SUM,
                returns_total=RETURNS_SUM,
                amount_total=Sum("amount"),
                quantity_total=Sum("quantity"),
            )
            .order_by("-amount_total")
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Продажи по товарам"
        sheet.append(
            [
                "Номенклатура",
                "Сумма продаж",
                "Возвраты и корректировки",
                "Чистая сумма",
                "Количество",
                "Средняя цена",
                "Доля в продажах",
            ]
        )

        for row in rows:
            amount = float(row["amount_total"] or 0)
            quantity = float(row["quantity_total"] or 0)
            sheet.append(
                [
                    row["nomenclature"],
                    float(row["gross_sales_total"] or 0),
                    float(row["returns_total"] or 0),
                    amount,
                    quantity,
                    amount / quantity if quantity else None,
                    amount / float(total_amount) if total_amount else None,
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sales_products.xlsx"'
        return response


class SalesCounterpartiesExportView(APIView):
    def get(self, request):
        level = request.query_params.get("level", "legal_entity")
        if level not in {"legal_entity", "brand", "counterparty_raw", "contract_number"}:
            level = "legal_entity"

        qs = _filtered_queryset(request)
        total_amount = qs.aggregate(total=Sum("amount"))["total"] or 0
        rows = (
            qs.values(level)
            .annotate(
                gross_sales_total=GROSS_SALES_SUM,
                returns_total=RETURNS_SUM,
                amount_total=Sum("amount"),
                quantity_total=Sum("quantity"),
                documents_count=Count("sales_doc_number", distinct=True),
            )
            .order_by("-amount_total")
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Продажи по контрагентам"
        sheet.append(
            [
                "Контрагент",
                "Сумма продаж",
                "Возвраты и корректировки",
                "Чистая сумма",
                "Количество",
                "Документов",
                "Доля в продажах",
            ]
        )

        for row in rows:
            amount = float(row["amount_total"] or 0)
            sheet.append(
                [
                    row[level],
                    float(row["gross_sales_total"] or 0),
                    float(row["returns_total"] or 0),
                    amount,
                    float(row["quantity_total"] or 0),
                    row["documents_count"] or 0,
                    amount / float(total_amount) if total_amount else None,
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sales_counterparties.xlsx"'
        return response
